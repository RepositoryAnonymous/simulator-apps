import sys
import copy
import os
import pickle

from ncu_report_helper import *

def find_nsight_folder(cuda_base_path):
    for root, dirs, files in os.walk(cuda_base_path):
        if "nsight-compute-" in root and "extras/python" in root:
            return root
    return None

cuda_base_path = "/usr/local/cuda-"
cuda_versions = ["11.0", "11.8", "12.0", "12.1", "12.3"]

ncu_report_path = None

for version in cuda_versions:
    cuda_path = cuda_base_path + version
    result = find_nsight_folder(cuda_path)
    if result:
        ncu_report_path = result
        print(f"\033[0;32;40mUse ncu_report.py moudule in path：{result}.\033[0m")
        break
else:
    print(f"\033[0;31;40mNot found version of cuda that enable ncu_report.py.\033[0m")
    exit()

sys.path.append(ncu_report_path)
import ncu_report

import openpyxl

import re

def parse_ppt_result_file(filename):
    with open(filename, 'r') as f:
        content = f.read()

    result = [None for _ in range(36)]
    result[ 0] = float(re.search(r'allocated max active thread block\(s\): (\d+)', content).group(1))
    result[ 1] = float(re.search(r'allocated max active warps per thread block: (\d+)', content).group(1))
    result[ 2] = int(re.search(r'Thread block Limit SM: (\d+)', content).group(1))
    result[ 3] = int(re.search(r'Thread block limit registers: (\d+)', content).group(1))
    result[ 4] = int(re.search(r'Thread block limit shared memory: (\d+)', content).group(1))
    result[ 5] = int(re.search(r'Thread block limit warps: (\d+)', content).group(1))
    result[ 6] = int(re.search(r'theoretical max active warps per SM: (\d+)', content).group(1))
    result[ 7] = float(re.search(r'theoretical occupancy: (\d+\.\d+|\d+) %', content).group(1))
    result[ 8] = float(re.search(r'achieved active warps per SM: (\d+\.\d+|\d+)', content).group(1))
    result[ 9] = float(re.search(r'achieved occupancy: (\d+\.\d+|\d+) %', content).group(1))
    result[10] = float(re.search(r'unified L1 cache hit rate: (\d+\.\d+|\d+) %', content).group(1))
    result[11] = float(re.search( \
        r'unified L1 cache hit rate for read transactions \(global memory accesses\): (\d+\.\d+|\d+) %', \
        content).group(1))
    result[12] = float(re.search(r'L2 cache hit rate: (\d+\.\d+|\d+) %', content).group(1))
    result[13] = int(re.search(r'GMEM read requests: (\d+)', content).group(1))
    result[14] = int(re.search(r'GMEM write requests: (\d+)', content).group(1))
    result[15] = int(re.search(r'GMEM total requests: (\d+)', content).group(1))
    result[16] = int(re.search(r'GMEM read transactions: (\d+)', content).group(1))
    result[17] = int(re.search(r'GMEM write transactions: (\d+)', content).group(1))
    result[18] = int(re.search(r'GMEM total transactions: (\d+)', content).group(1))
    result[19] = float(re.search(r'number of read transactions per read requests: (\d+\.\d+|\d+) \(', \
                       content).group(1))
    result[20] = float(re.search(r'number of write transactions per write requests: (\d+\.\d+|\d+) \(', \
                                 content).group(1))
    result[21] = int(re.search(r'L2 read transactions: (\d+)', content).group(1))
    result[22] = int(re.search(r'L2 write transactions: (\d+)', content).group(1))
    result[23] = int(re.search(r'L2 total transactions: (\d+)', content).group(1))
    result[24] = int(re.search(r'DRAM total transactions: (\d+)', content).group(1))
    result[25] = int(re.search(r'Total number of global atomic requests: (\d+)', content).group(1))
    result[26] = int(re.search(r'Total number of global reduction requests: (\d+)', content).group(1))
    result[27] = int(re.search(r'Global memory atomic and reduction transactions: (\d+)', content).group(1))
    result[28] = int(re.search(r'GPU active cycles \(max\): (\d{1,3}(,\d{3})*|\d+)', \
        content).group(1).replace(",", ""))
    result[29] = int(re.search(r'SM active cycles \(sum\): (\d{1,3}(,\d{3})*|\d+)', \
        content).group(1).replace(",", ""))
    result[30] = int(re.search(r'Warp instructions executed: (\d{1,3}(,\d{3})*|\d+)', \
        content).group(1).replace(",", ""))
    result[31] = float(re.search(r'Instructions executed per clock cycle \(IPC\): (\d+\.\d+|\d+)', \
        content).group(1))
    result[32] = float(re.search(r'Total instructions executed per seconds \(MIPS\): (\d+\.\d+|\d+)', \
        content).group(1))
    result[33] = float(re.search(r'Kernel execution time: (\d+\.\d+|\d+)', content).group(1))
    if re.search(r'Kernel execution time: (\d+\.\d+|\d+) ([a-zA-Z]+)', content).group(2) == "us":
        result[33] = result[33]*1000
    result[34] = float(re.search(r'Memory model: (\d+\.\d+) sec, \d+:\d+:\d+', content).group(1))
    result[35] = float(re.search(r'Compute model: (\d+\.\d+) sec, \d+:\d+:\d+', content).group(1))

    return result

def find_row_num(file_path, string, num_occur):
    with open(file_path, 'r') as file:
        lines = file.readlines()
        count = 0
        for index, line in enumerate(lines):
            if string in line:
                count += 1
                if count == num_occur:
                    return index + 1
    return -1

def parse_asim_result_file(filename, start, end):
    with open(filename, 'r') as f:
        content_all = f.read()
        
        content_ = content_all.split('\n')[start:end+1]
        content = ""
        for char in content_:
            content += char + '\n'

    result = [None for _ in range(36)]
    result[ 0] = None
    result[ 1] = None
    result[ 2] = None
    result[ 3] = None
    result[ 4] = None
    result[ 5] = None
    result[ 6] = None
    result[ 7] = None
    result[ 8] = None
    result[ 9] = float(re.search(r'gpu_occupancy = (\d+\.\d+|\d+)%', content).group(1)) 
    result[10] = (1. - float(re.search(r'L1D_total_cache_miss_rate = (\d+\.\d+|\d+)', content).group(1))) * 100
    result[11] = None
    result[12] = (1. - float(re.search(r'L2_total_cache_miss_rate = (\d+\.\d+|\d+)', content).group(1))) * 100
    result[13] = int(re.search(r'gpgpu_n_mem_read_global = (\d+)', content).group(1))
    result[14] = int(re.search(r'gpgpu_n_mem_write_global = (\d+)', content).group(1)) 
    result[15] = int(re.search(r'gpgpu_n_mem_read_global = (\d+)', content).group(1)) + \
                 int(re.search(r'gpgpu_n_mem_write_global = (\d+)', content).group(1))
    result[16] = None
    result[17] = None
    result[18] = None
    result[19] = None
    result[20] = None
    result[21] = None
    result[22] = None
    result[23] = None
    result[24] = None
    result[25] = None
    result[26] = None
    result[27] = None
    result[28] = int(re.search(r'gpu_sim_cycle = (\d+)', content).group(1))
    result[29] = int(re.search(r'gpu_sim_cycle = (\d+)', content).group(1))
    result[30] = int(int(re.search(r'gpu_sim_insn = (\d+)', content).group(1)) / 32)
    result[31] = float(re.search(r'gpu_ipc =\s*(\d+\.\d+|\d+)', content).group(1)) / 32 / \
                 float(re.search(r'-gpgpu_occupancy_sm_number\s*(\d+)', content_all).group(1))
    result[32] = result[31] * float(re.search(r'-gpgpu_clock_domains (\d+\.\d+|\d+):', content_all).group(1))
    result[33] = float(re.search(r'gpgpu_simulation_rate = (\d+) \(cycle', content).group(1)) * \
                       int(re.search( \
                           r'gpgpu_simulation_time = (\d+) days, (\d+) hrs, (\d+) min, (\d+) sec \((\d+) sec\)', \
                           content).group(5)) / \
                       float(re.search(r'-gpgpu_clock_domains\s*(\d+\.\d+|\d+):', content_all).group(1)) * 1e3
    result[34] = int(re.search(\
                     r'gpgpu_simulation_time = (\d+) days, (\d+) hrs, (\d+) min, (\d+) sec \((\d+) sec\)', \
                     content).group(5))
    result[35] = None

    return result

def parse_ours_result_file(filename):
    # judge if the file exists, if not, return a list of None
    if not os.path.exists(filename):
        return [None for _ in range(36)], [None for _ in range(108)]
    
    with open(filename, 'r') as f:
        content = f.read()
    
    result = [None for _ in range(36)]
    result[ 0] = None
    result[ 1] = None
    result[ 2] = int(re.search(r'Thread_block_limit_SM: (\d+)', content).group(1))
    result[ 3] = int(re.search(r'Thread_block_limit_registers: (\d+)', content).group(1))
    result[ 4] = int(re.search(r'Thread_block_limit_shared_memory: (\d+)', content).group(1))
    result[ 5] = int(re.search(r'Thread_block_limit_warps: (\d+)', content).group(1))
    result[ 6] = int(re.search(r'Theoretical_max_active_warps_per_SM: (\d+)', content).group(1))
    result[ 7] = float(re.search(r'Theoretical_occupancy: (\d+\.\d+|\d+)', content).group(1)) * 100.0
    result[ 8] = float(re.search(r'Achieved_active_warps_per_SM: (\d+\.\d+|\d+)', content).group(1))
    result[ 9] = float(re.search(r'Achieved_occupancy: (\d+\.\d+|\d+)', content).group(1)) * 100.0
    result[10] = float(re.search(r'Unified_L1_cache_hit_rate: (\d+\.\d+|\d+)', content).group(1)) * 100.0
    result[11] = None
    result[12] = float(re.search(r'L2_cache_hit_rate: (\d+\.\d+|\d+)', content).group(1)) * 100.0
    result[13] = int(re.search(r'GMEM_read_requests: (\d+)', content).group(1))
    result[14] = int(re.search(r'GMEM_write_requests: (\d+)', content).group(1))
    result[15] = int(re.search(r'GMEM_total_requests: (\d+)', content).group(1))
    result[16] = int(re.search(r'GMEM_read_transactions: (\d+)', content).group(1))
    result[17] = int(re.search(r'GMEM_write_transactions: (\d+)', content).group(1))
    result[18] = int(re.search(r'GMEM_total_transactions: (\d+)', content).group(1))
    result[19] = float(re.search(r'Number_of_read_transactions_per_read_requests: (\d+\.\d+|\d+)', \
                                 content).group(1))
    result[20] = float(re.search(r'Number_of_write_transactions_per_write_requests: (\d+\.\d+|\d+)', \
                                 content).group(1))
    result[21] = int(re.search(r'L2_read_transactions: (\d+)', content).group(1))
    result[22] = int(re.search(r'L2_write_transactions: (\d+)', content).group(1))
    result[23] = int(re.search(r'L2_total_transactions: (\d+)', content).group(1))
    result[24] = int(re.search(r'DRAM_total_transactions: (\d+)', content).group(1))
    result[25] = int(re.search(r'Total_number_of_global_atomic_requests: (\d+)', content).group(1))
    result[26] = int(re.search(r'Total_number_of_global_reduction_requests: (\d+)', content).group(1))
    result[27] = int(re.search(r'Global_memory_atomic_and_reduction_transactions: (\d+)', content).group(1))
    result[28] = int(re.search(r'GPU_active_cycles: (\d+)', content).group(1).replace(",", ""))
    result[29] = int(re.search(r'SM_active_cycles: (\d+)', content).group(1).replace(",", ""))
    
    result[30] = int(re.search(r'Warp_instructions_executed: (\d+)', content).group(1).replace(",", ""))
    result[31] = float(re.search(r'Instructions_executed_per_clock_cycle_IPC: (\d+\.\d+|\d+)', \
                                 content).group(1))
    result[32] = float(re.search(r'Total_instructions_executed_per_seconds \(MIPS\): (\d+\.\d+|\d+)', \
                                 content).group(1))
    result[33] = float(re.search(r'Kernel_execution_time \(ns\): (\d+\.\d+|\d+)', content).group(1))
    result[34] = float(re.search(r'Simulation_time_memory_model \(s\): (\d+\.\d+|\d+)', content).group(1))
    result[35] = float(re.search(r'Simulation_time_compute_model \(s\): (\d+\.\d+|\d+)', content).group(1))
    
    result_stall = [None for _ in range(108)]
    
    result_stall[0] = int(re.search(r'Compute_Structural_Stall: (\d+)', content).group(1))
    result_stall[1] = int(re.search(r'Compute_Data_Stall: (\d+)', content).group(1))
    result_stall[2] = int(re.search(r'Memory_Structural_Stall: (\d+)', content).group(1))
    result_stall[3] = int(re.search(r'Memory_Data_Stall: (\d+)', content).group(1))
    result_stall[4] = int(re.search(r'Synchronization_Stall: (\d+)', content).group(1))
    result_stall[5] = int(re.search(r'Control_Stall: (\d+)', content).group(1))
    result_stall[6] = int(re.search(r'Idle_Stall: (\d+)', content).group(1))
    result_stall[7] = int(re.search(r'Other_Stall: (\d+)', content).group(1))
    result_stall[8] = int(re.search(r'No_Stall: (\d+)', content).group(1))
    
    result_stall[9] = float(re.search(r'Compute_Structural_Stall_ratio: (\d+\.\d+|\d+)', content).group(1))
    result_stall[10] = float(re.search(r'Compute_Data_Stall_ratio: (\d+\.\d+|\d+)', content).group(1))
    result_stall[11] = float(re.search(r'Memory_Structural_Stall_ratio: (\d+\.\d+|\d+)', content).group(1))
    result_stall[12] = float(re.search(r'Memory_Data_Stall_ratio: (\d+\.\d+|\d+)', content).group(1))
    result_stall[13] = float(re.search(r'Synchronization_Stall_ratio: (\d+\.\d+|\d+)', content).group(1))
    result_stall[14] = float(re.search(r'Control_Stall_ratio: (\d+\.\d+|\d+)', content).group(1))
    result_stall[15] = float(re.search(r'Idle_Stall_ratio: (\d+\.\d+|\d+)', content).group(1))
    result_stall[16] = float(re.search(r'Other_Stall_ratio: (\d+\.\d+|\d+)', content).group(1))
    result_stall[17] = float(re.search(r'No_Stall_ratio: (\d+\.\d+|\d+)', content).group(1))
    
    MemoryStructuralStallCyclesBreakdownMatch = \
        re.search(r'Memory Structural Stall Cycles Breakdown:', content)
    MemoryStructuralStallCyclesBreakdownDistributionMatch = \
        re.search(r'Memory Structural Stall Cycles Breakdown Distribution:', content)
    ComputeStructuralStallCyclesBreakdownMatch = \
        re.search(r'Compute Structural Stall Cycles Breakdown:', content)
    ComputeStructuralStallCyclesBreakdownDistributionMatch = \
        re.search(r'Compute Structural Stall Cycles Breakdown Distribution:', content)

    result_stall[18] = float(re.search(r'Issue_out_has_no_free_slot: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownMatch.end():\
            MemoryStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[19] = float(re.search(r'Issue_previous_issued_inst_exec_type_is_memory: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownMatch.end():\
            MemoryStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[20] = float(re.search(r'Execute_result_bus_has_no_slot_for_latency: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownMatch.end():\
            MemoryStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[21] = float(re.search(r'Execute_m_dispatch_reg_of_fu_is_not_empty: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownMatch.end():\
            MemoryStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[22] = float(re.search(r'Writeback_bank_of_reg_is_not_idle: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownMatch.end():\
            MemoryStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[23] = float(re.search(r'ReadOperands_bank_reg_belonged_to_was_allocated: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownMatch.end():\
            MemoryStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[24] = float(re.search(\
        r'ReadOperands_port_num_m_in_ports_m_in_fails_as_not_found_free_cu: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownMatch.end():\
            MemoryStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[25] = float(re.search(r'Execute_icnt_injection_buffer_is_full: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownMatch.end():\
            MemoryStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))

    result_stall[26] = float(re.search(r'Issue_out_has_no_free_slot: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownDistributionMatch.end():\
            ComputeStructuralStallCyclesBreakdownMatch.start()+1]).group(1))
    result_stall[27] = float(re.search(r'Issue_previous_issued_inst_exec_type_is_memory: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownDistributionMatch.end():\
            ComputeStructuralStallCyclesBreakdownMatch.start()+1]).group(1))
    result_stall[28] = float(re.search(r'Execute_result_bus_has_no_slot_for_latency: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownDistributionMatch.end():\
            ComputeStructuralStallCyclesBreakdownMatch.start()+1]).group(1))
    result_stall[29] = float(re.search(r'Execute_m_dispatch_reg_of_fu_is_not_empty: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownDistributionMatch.end():\
            ComputeStructuralStallCyclesBreakdownMatch.start()+1]).group(1))
    result_stall[30] = float(re.search(r'Writeback_bank_of_reg_is_not_idle: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownDistributionMatch.end():\
            ComputeStructuralStallCyclesBreakdownMatch.start()+1]).group(1))
    result_stall[31] = float(re.search(r'ReadOperands_bank_reg_belonged_to_was_allocated: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownDistributionMatch.end():\
            ComputeStructuralStallCyclesBreakdownMatch.start()+1]).group(1))
    result_stall[32] = float(re.search(\
        r'ReadOperands_port_num_m_in_ports_m_in_fails_as_not_found_free_cu: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownDistributionMatch.end():\
            ComputeStructuralStallCyclesBreakdownMatch.start()+1]).group(1))
    result_stall[33] = float(re.search(r'Execute_icnt_injection_buffer_is_full: (\d+\.\d+|\d+)', \
        content[MemoryStructuralStallCyclesBreakdownDistributionMatch.end():\
            ComputeStructuralStallCyclesBreakdownMatch.start()+1]).group(1))

    result_stall[34] = float(re.search(r'Issue_out_has_no_free_slot: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownMatch.end():\
            ComputeStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[35] = float(re.search(r'Issue_previous_issued_inst_exec_type_is_compute: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownMatch.end():\
            ComputeStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[36] = float(re.search(r'Execute_result_bus_has_no_slot_for_latency: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownMatch.end():\
            ComputeStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[37] = float(re.search(r'Execute_m_dispatch_reg_of_fu_is_not_empty: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownMatch.end():\
            ComputeStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[38] = float(re.search(r'Writeback_bank_of_reg_is_not_idle: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownMatch.end():\
            ComputeStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[39] = float(re.search(r'ReadOperands_bank_reg_belonged_to_was_allocated: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownMatch.end():\
            ComputeStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[40] = float(re.search(\
        r'ReadOperands_port_num_m_in_ports_m_in_fails_as_not_found_free_cu: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownMatch.end():\
            ComputeStructuralStallCyclesBreakdownDistributionMatch.start()+1]).group(1))

    result_stall[41] = float(re.search(r'Issue_out_has_no_free_slot: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[42] = float(re.search(r'Issue_previous_issued_inst_exec_type_is_compute: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[43] = float(re.search(r'Execute_result_bus_has_no_slot_for_latency: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[44] = float(re.search(r'Execute_m_dispatch_reg_of_fu_is_not_empty: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[45] = float(re.search(r'Writeback_bank_of_reg_is_not_idle: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[46] = float(re.search(r'ReadOperands_bank_reg_belonged_to_was_allocated: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[47] = float(re.search(\
        r'ReadOperands_port_num_m_in_ports_m_in_fails_as_not_found_free_cu: (\d+\.\d+|\d+)', \
        content[ComputeStructuralStallCyclesBreakdownDistributionMatch.end():]).group(1))

    MemoryDataStallCyclesBreakdownMatch = \
        re.search(r'Memory Data Stall Cycles Breakdown:', content)
    MemoryDataStallCyclesBreakdownDistributionMatch = \
        re.search(r'Memory Data Stall Cycles Breakdown Distribution:', content)
    ComputeDataStallCyclesBreakdownMatch = \
        re.search(r'Compute Data Stall Cycles Breakdown:', content)
    ComputeDataStallCyclesBreakdownDistributionMatch = \
        re.search(r'Compute Data Stall Cycles Breakdown Distribution:', content)
        
    result_stall[48] = float(re.search(\
        r'Issue_scoreboard: (\d+\.\d+|\d+)', \
        content[MemoryDataStallCyclesBreakdownMatch.end():\
            MemoryDataStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[49] = float(re.search(\
        r'Execute_L1: (\d+\.\d+|\d+)', \
        content[MemoryDataStallCyclesBreakdownMatch.end():\
            MemoryDataStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[50] = float(re.search(\
        r'Execute_L2: (\d+\.\d+|\d+)', \
        content[MemoryDataStallCyclesBreakdownMatch.end():\
            MemoryDataStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[51] = float(re.search(\
        r'Execute_Main_Memory: (\d+\.\d+|\d+)', \
        content[MemoryDataStallCyclesBreakdownMatch.end():\
            MemoryDataStallCyclesBreakdownDistributionMatch.start()+1]).group(1))

    result_stall[52] = float(re.search(\
        r'Issue_scoreboard: (\d+\.\d+|\d+)', \
        content[MemoryDataStallCyclesBreakdownDistributionMatch.end():\
            ComputeDataStallCyclesBreakdownMatch.start()+1]).group(1))
    result_stall[53] = float(re.search(\
        r'Execute_L1: (\d+\.\d+|\d+)', \
        content[MemoryDataStallCyclesBreakdownDistributionMatch.end():\
            ComputeDataStallCyclesBreakdownMatch.start()+1]).group(1))
    result_stall[54] = float(re.search(\
        r'Execute_L2: (\d+\.\d+|\d+)', \
        content[MemoryDataStallCyclesBreakdownDistributionMatch.end():\
            ComputeDataStallCyclesBreakdownMatch.start()+1]).group(1))
    result_stall[55] = float(re.search(\
        r'Execute_Main_Memory: (\d+\.\d+|\d+)', \
        content[MemoryDataStallCyclesBreakdownDistributionMatch.end():\
            ComputeDataStallCyclesBreakdownMatch.start()+1]).group(1))

    result_stall[56] = float(re.search(\
        r'Issue_scoreboard: (\d+\.\d+|\d+)', \
        content[ComputeDataStallCyclesBreakdownMatch.end():\
            ComputeDataStallCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[57] = float(re.search(\
        r'Issue_scoreboard: (\d+\.\d+|\d+)', \
        content[ComputeDataStallCyclesBreakdownDistributionMatch.end():]).group(1))

    FunctionUnitExecutionCyclesBreakdownMatch = \
        re.search(r'Function Unit Execution Cycles Breakdown:', content)
    FunctionUnitExecutionCyclesBreakdownDistributionMatch = \
        re.search(r'Function Unit Execution Cycles Breakdown Distribution:', content)
        
    result_stall[58] = float(re.search(\
        r'SP_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownMatch.end():\
            FunctionUnitExecutionCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[59] = float(re.search(\
        r'SFU_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownMatch.end():\
            FunctionUnitExecutionCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[60] = float(re.search(\
        r'INT_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownMatch.end():\
            FunctionUnitExecutionCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[61] = float(re.search(\
        r'DP_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownMatch.end():\
            FunctionUnitExecutionCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[62] = float(re.search(\
        r'TENSOR_CORE_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownMatch.end():\
            FunctionUnitExecutionCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[63] = float(re.search(\
        r'LDST_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownMatch.end():\
            FunctionUnitExecutionCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[64] = float(re.search(\
        r'SPEC_UNIT_1_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownMatch.end():\
            FunctionUnitExecutionCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[65] = float(re.search(\
        r'SPEC_UNIT_2_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownMatch.end():\
            FunctionUnitExecutionCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[66] = float(re.search(\
        r'SPEC_UNIT_3_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownMatch.end():\
            FunctionUnitExecutionCyclesBreakdownDistributionMatch.start()+1]).group(1))
    result_stall[67] = float(re.search(\
        r'Other_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownMatch.end():\
            FunctionUnitExecutionCyclesBreakdownDistributionMatch.start()+1]).group(1))

    result_stall[68] = float(re.search(\
        r'SP_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[69] = float(re.search(\
        r'SFU_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[70] = float(re.search(\
        r'INT_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[71] = float(re.search(\
        r'DP_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[72] = float(re.search(\
        r'TENSOR_CORE_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[73] = float(re.search(\
        r'LDST_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[74] = float(re.search(\
        r'SPEC_UNIT_1_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[75] = float(re.search(\
        r'SPEC_UNIT_2_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[76] = float(re.search(\
        r'SPEC_UNIT_3_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownDistributionMatch.end():]).group(1))
    result_stall[77] = float(re.search(\
        r'Other_UNIT_execute_clks: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionCyclesBreakdownDistributionMatch.end():]).group(1))
  
    FunctionUnitExecutionInstnsNumberMatch = \
        re.search(r'Function Unit Execution Instns Number:', content)
    FunctionUnitExecutionInstnsNumberDistributionMatch = \
        re.search(r'Function Unit Execution Instns Number Distribution:', content)
        
    result_stall[78] = float(re.search(\
        r'SP_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberMatch.end():\
            FunctionUnitExecutionInstnsNumberDistributionMatch.start()+1]).group(1))
    result_stall[79] = float(re.search(\
        r'SFU_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberMatch.end():\
            FunctionUnitExecutionInstnsNumberDistributionMatch.start()+1]).group(1))
    result_stall[80] = float(re.search(\
        r'INT_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberMatch.end():\
            FunctionUnitExecutionInstnsNumberDistributionMatch.start()+1]).group(1))
    result_stall[81] = float(re.search(\
        r'DP_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberMatch.end():\
            FunctionUnitExecutionInstnsNumberDistributionMatch.start()+1]).group(1))
    result_stall[82] = float(re.search(\
        r'TENSOR_CORE_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberMatch.end():\
            FunctionUnitExecutionInstnsNumberDistributionMatch.start()+1]).group(1))
    result_stall[83] = float(re.search(\
        r'LDST_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberMatch.end():\
            FunctionUnitExecutionInstnsNumberDistributionMatch.start()+1]).group(1))
    result_stall[84] = float(re.search(\
        r'SPEC_UNIT_1_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberMatch.end():\
            FunctionUnitExecutionInstnsNumberDistributionMatch.start()+1]).group(1))
    result_stall[85] = float(re.search(\
        r'SPEC_UNIT_2_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberMatch.end():\
            FunctionUnitExecutionInstnsNumberDistributionMatch.start()+1]).group(1))
    result_stall[86] = float(re.search(\
        r'SPEC_UNIT_3_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberMatch.end():\
            FunctionUnitExecutionInstnsNumberDistributionMatch.start()+1]).group(1))
    result_stall[87] = float(re.search(\
        r'Other_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberMatch.end():\
            FunctionUnitExecutionInstnsNumberDistributionMatch.start()+1]).group(1))

    result_stall[88] = float(re.search(\
        r'SP_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberDistributionMatch.end():]).group(1))
    result_stall[89] = float(re.search(\
        r'SFU_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberDistributionMatch.end():]).group(1))
    result_stall[90] = float(re.search(\
        r'INT_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberDistributionMatch.end():]).group(1))
    result_stall[91] = float(re.search(\
        r'DP_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberDistributionMatch.end():]).group(1))
    result_stall[92] = float(re.search(\
        r'TENSOR_CORE_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberDistributionMatch.end():]).group(1))
    result_stall[93] = float(re.search(\
        r'LDST_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberDistributionMatch.end():]).group(1))
    result_stall[94] = float(re.search(\
        r'SPEC_UNIT_1_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberDistributionMatch.end():]).group(1))
    result_stall[95] = float(re.search(\
        r'SPEC_UNIT_2_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberDistributionMatch.end():]).group(1))
    result_stall[96] = float(re.search(\
        r'SPEC_UNIT_3_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberDistributionMatch.end():]).group(1))
    result_stall[97] = float(re.search(\
        r'Other_UNIT_Instns_num: (\d+\.\d+|\d+)', \
        content[FunctionUnitExecutionInstnsNumberDistributionMatch.end():]).group(1))

    result_stall[98] = float(re.search(\
        r'SP_UNIT_average_cycles_per_instn: (\d+\.\d+|\d+)', content).group(1))
    result_stall[99] = float(re.search(\
        r'SFU_UNIT_average_cycles_per_instn: (\d+\.\d+|\d+)', content).group(1))
    result_stall[100] = float(re.search(\
        r'INT_UNIT_average_cycles_per_instn: (\d+\.\d+|\d+)', content).group(1))
    result_stall[101] = float(re.search(\
        r'DP_UNIT_average_cycles_per_instn: (\d+\.\d+|\d+)', content).group(1))
    result_stall[102] = float(re.search(\
        r'TENSOR_CORE_UNIT_average_cycles_per_instn: (\d+\.\d+|\d+)', content).group(1))
    result_stall[103] = float(re.search(\
        r'LDST_UNIT_average_cycles_per_instn: (\d+\.\d+|\d+)', content).group(1))
    result_stall[104] = float(re.search(\
        r'SPEC_UNIT_1_average_cycles_per_instn: (\d+\.\d+|\d+)', content).group(1))
    result_stall[105] = float(re.search(\
        r'SPEC_UNIT_2_average_cycles_per_instn: (\d+\.\d+|\d+)', content).group(1))
    result_stall[106] = float(re.search(\
        r'SPEC_UNIT_3_average_cycles_per_instn: (\d+\.\d+|\d+)', content).group(1))
    result_stall[107] = float(re.search(\
        r'Other_UNIT_average_cycles_per_instn: (\d+\.\d+|\d+)', content).group(1))

    return result, result_stall

if __name__ == "__main__":
    if os.path.exists('compare.xlsx'):
        workbook = openpyxl.load_workbook('compare.xlsx')
    else:
        workbook = openpyxl.Workbook()

    sheet_ncu = workbook.create_sheet('NCU', 1)
    sheet_ppt = workbook.create_sheet('PPT', 2)
    err_ppt = workbook.create_sheet('ERR-PPT', 3)
    sheet_accel_sim = workbook.create_sheet('ASIM', 4)
    err_accel_sim = workbook.create_sheet('ERR-ASIM', 5)
    sheet_ours = workbook.create_sheet('OURS', 6)
    err_ours = workbook.create_sheet('ERR-OURS', 7)

    entry = [ \
        "",\
        "Kernel ID", \
        "Thread block Limit SM", \
        "Thread block limit registers", \
        "Thread block limit shared memory", \
        "Thread block limit warps", \
        "theoretical max active warps per SM", \
        "theoretical occupancy", \
        "achieved active warps per SM", \
        "achieved occupancy", \
        "unified L1 cache hit rate", \
        "unified L1 cache hit rate for read transactions (global memory accesses)", \
        "L2 cache hit rate", \
        "GMEM read requests", \
        "GMEM write requests", \
        "GMEM total requests", \
        "GMEM read transactions", \
        "GMEM write transactions", \
        "GMEM total transactions", \
        "number of read transactions per read requests", \
        "number of write transactions per write requests", \
        "L2 read transactions", \
        "L2 write transactions", \
        "L2 total transactions", \
        "DRAM total transactions", \
        "Total number of global atomic requests", \
        "Total number of global reduction requests", \
        "Global memory atomic and reduction transactions", \
        "GPU active cycles", \
        "SM active cycles", \
        "Warp instructions executed", \
        "Instructions executed per clock cycle (IPC)", \
        "Total instructions executed per seconds (MIPS)", \
        "Kernel execution time (ns)", \
    ]
    entry_ours = entry + [
        "Memory model time (s)",
        "Compute model time (s)",
        "Compute_Structural_Stall_Cycles",
        "Compute_Data_Stall_Cycles",
        "Memory_Structural_Stall_Cycles",
        "Memory_Data_Stall_Cycles",
        "Synchronization_Stall_Cycles",
        "Control_Stall_Cycles",
        "Idle_Stall_Cycles",
        "Other_Stall_Cycles",
        "No_Stall_Cycles",
        "Compute_Structural_Stall_Ratio",
        "Compute_Data_Stall_Ratio",
        "Memory_Structural_Stall_Ratio",
        "Memory_Data_Stall_Ratio",
        "Synchronization_Stall_Ratio",
        "Control_Stall_Ratio",
        "Idle_Stall_Ratio",
        "Other_Stall_Ratio",
        "No_Stall_Ratio",
        "MemoryStructuralStall_Issue_out_has_no_free_slot_Cycles",
        "MemoryStructuralStall_Issue_previous_issued_inst_exec_type_is_memory_Cycles",
        "MemoryStructuralStall_Execute_result_bus_has_no_slot_for_latency_Cycles",
        "MemoryStructuralStall_Execute_m_dispatch_reg_of_fu_is_not_empty_Cycles",
        "MemoryStructuralStall_Writeback_bank_of_reg_is_not_idle_Cycles",
        "MemoryStructuralStall_ReadOperands_bank_reg_belonged_to_was_allocated_Cycles",
        "MemoryStructuralStall_ReadOperands_port_num_m_in_ports_m_in_fails_as_not_found_free_cu_Cycles",
        "MemoryStructuralStall_Execute_icnt_injection_buffer_is_full_Cycles",
        "MemoryStructuralStall_Issue_out_has_no_free_slot_Ratio",
        "MemoryStructuralStall_Issue_previous_issued_inst_exec_type_is_memory_Ratio",
        "MemoryStructuralStall_Execute_result_bus_has_no_slot_for_latency_Ratio",
        "MemoryStructuralStall_Execute_m_dispatch_reg_of_fu_is_not_empty_Ratio",
        "MemoryStructuralStall_Writeback_bank_of_reg_is_not_idle_Ratio",
        "MemoryStructuralStall_ReadOperands_bank_reg_belonged_to_was_allocated_Ratio",
        "MemoryStructuralStall_ReadOperands_port_num_m_in_ports_m_in_fails_as_not_found_free_cu_Ratio",
        "MemoryStructuralStall_Execute_icnt_injection_buffer_is_full_Ratio",
        "ComputeStructuralStall_Issue_out_has_no_free_slot_Cycles",
        "ComputeStructuralStall_Issue_previous_issued_inst_exec_type_is_compute_Cycles",
        "ComputeStructuralStall_Execute_result_bus_has_no_slot_for_latency_Cycles",
        "ComputeStructuralStall_Execute_m_dispatch_reg_of_fu_is_not_empty_Cycles",
        "ComputeStructuralStall_Writeback_bank_of_reg_is_not_idle_Cycles",
        "ComputeStructuralStall_ReadOperands_bank_reg_belonged_to_was_allocated_Cycles",
        "ComputeStructuralStall_ReadOperands_port_num_m_in_ports_m_in_fails_as_not_found_free_cu_Cycles",
        "ComputeStructuralStall_Issue_out_has_no_free_slot_Ratio",
        "ComputeStructuralStall_Issue_previous_issued_inst_exec_type_is_compute_Ratio",
        "ComputeStructuralStall_Execute_result_bus_has_no_slot_for_latency_Ratio",
        "ComputeStructuralStall_Execute_m_dispatch_reg_of_fu_is_not_empty_Ratio",
        "ComputeStructuralStall_Writeback_bank_of_reg_is_not_idle_Ratio",
        "ComputeStructuralStall_ReadOperands_bank_reg_belonged_to_was_allocated_Ratio",
        "ComputeStructuralStall_ReadOperands_port_num_m_in_ports_m_in_fails_as_not_found_free_cu_Ratio",
        "MemoryDataStall_Issue_scoreboard_Cycles",
        "MemoryDataStall_Execute_L1_Cycles",
        "MemoryDataStall_Execute_L2_Cycles",
        "MemoryDataStall_Execute_Main_Memory_Cycles",
        "MemoryDataStall_Issue_scoreboard_Ratio",
        "MemoryDataStall_Execute_L1_Ratio",
        "MemoryDataStall_Execute_L2_Ratio",
        "MemoryDataStall_Execute_Main_Memory_Ratio",
        "ComputeDataStall_Issue_scoreboard_Cycles",
        "ComputeDataStall_Issue_scoreboard_Ratio",
        "FunctionUnitExecution_SP_UNIT_Cycles",
        "FunctionUnitExecution_SFU_UNIT_Cycles",
        "FunctionUnitExecution_INT_UNIT_Cycles",
        "FunctionUnitExecution_DP_UNIT_Cycles",
        "FunctionUnitExecution_TENSOR_CORE_UNIT_Cycles",
        "FunctionUnitExecution_LDST_UNIT_Cycles",
        "FunctionUnitExecution_SPEC_UNIT_1_Cycles",
        "FunctionUnitExecution_SPEC_UNIT_2_Cycles",
        "FunctionUnitExecution_SPEC_UNIT_3_Cycles",
        "FunctionUnitExecution_Other_UNIT_Cycles",
        "FunctionUnitExecution_SP_UNIT_Ratio",
        "FunctionUnitExecution_SFU_UNIT_Ratio",
        "FunctionUnitExecution_INT_UNIT_Ratio",
        "FunctionUnitExecution_DP_UNIT_Ratio",
        "FunctionUnitExecution_TENSOR_CORE_UNIT_Ratio",
        "FunctionUnitExecution_LDST_UNIT_Ratio",
        "FunctionUnitExecution_SPEC_UNIT_1_Ratio",
        "FunctionUnitExecution_SPEC_UNIT_2_Ratio",
        "FunctionUnitExecution_SPEC_UNIT_3_Ratio",
        "FunctionUnitExecution_Other_UNIT_Ratio",
        "FunctionUnitExecution_SP_UNIT_InstnsNumber",
        "FunctionUnitExecution_SFU_UNIT_InstnsNumber",
        "FunctionUnitExecution_INT_UNIT_InstnsNumber",
        "FunctionUnitExecution_DP_UNIT_InstnsNumber",
        "FunctionUnitExecution_TENSOR_CORE_UNIT_InstnsNumber",
        "FunctionUnitExecution_LDST_UNIT_InstnsNumber",
        "FunctionUnitExecution_SPEC_UNIT_1_InstnsNumber",
        "FunctionUnitExecution_SPEC_UNIT_2_InstnsNumber",
        "FunctionUnitExecution_SPEC_UNIT_3_InstnsNumber",
        "FunctionUnitExecution_Other_UNIT_InstnsNumber",
        "FunctionUnitExecution_SP_UNIT_InstnsNumber_Ratio",
        "FunctionUnitExecution_SFU_UNIT_InstnsNumber_Ratio",
        "FunctionUnitExecution_INT_UNIT_InstnsNumber_Ratio",
        "FunctionUnitExecution_DP_UNIT_InstnsNumber_Ratio",
        "FunctionUnitExecution_TENSOR_CORE_UNIT_InstnsNumber_Ratio",
        "FunctionUnitExecution_LDST_UNIT_InstnsNumber_Ratio",
        "FunctionUnitExecution_SPEC_UNIT_1_InstnsNumber_Ratio",
        "FunctionUnitExecution_SPEC_UNIT_2_InstnsNumber_Ratio",
        "FunctionUnitExecution_SPEC_UNIT_3_InstnsNumber_Ratio",
        "FunctionUnitExecution_Other_UNIT_InstnsNumber_Ratio",
        "FunctionUnitExecution_SP_UNIT_AverageCyclesPerInstn",
        "FunctionUnitExecution_SFU_UNIT_AverageCyclesPerInstn",
        "FunctionUnitExecution_INT_UNIT_AverageCyclesPerInstn",
        "FunctionUnitExecution_DP_UNIT_AverageCyclesPerInstn",
        "FunctionUnitExecution_TENSOR_CORE_UNIT_AverageCyclesPerInstn",
        "FunctionUnitExecution_LDST_UNIT_AverageCyclesPerInstn",
        "FunctionUnitExecution_SPEC_UNIT_1_AverageCyclesPerInstn",
        "FunctionUnitExecution_SPEC_UNIT_2_AverageCyclesPerInstn",
        "FunctionUnitExecution_SPEC_UNIT_3_AverageCyclesPerInstn",
        "FunctionUnitExecution_Other_UNIT_AverageCyclesPerInstn",
    ]
    
    sheet_ncu.append(entry + ["unified L1 cache total requests", "unified L2 cache total requests"])
    sheet_ppt.append(entry)
    err_ppt.append(entry)
    sheet_accel_sim.append(entry)
    err_accel_sim.append(entry)
    sheet_ours.append(entry_ours)
    err_ours.append(entry)

    for report_file_path in ncu_report_file_paths:
        report = ncu_report.load_report(report_file_path[0])
        kernel_num = min(len(report[0]), 100)

        for knum in range(kernel_num):
            kernel = report[0][knum]
        
        ################################################################################
        ####                             do ncu report                              ####
        ################################################################################
            print("%81s" % report_file_path[0].split("/")[-1], "knums: %4d" % \
                len(report[0]), "kernel-%4d" % (knum + 1))
            app_results = [report_file_path[0].split("/")[-1].split(".")[0], str(knum)]

            app_results.append(get_launch__occupancy_limit_blocks(kernel))
            app_results.append(get_launch__occupancy_limit_registers(kernel))
            app_results.append(get_launch__occupancy_limit_shared_mem(kernel))
            app_results.append(get_launch__occupancy_limit_warps(kernel))
            app_results.append(get_sm__maximum_warps_avg_per_active_cycle(kernel))
            app_results.append(get_sm__maximum_warps_per_active_cycle_pct(kernel))
            app_results.append(get_sm__warps_active_avg_per_cycle_active(kernel))
            app_results.append(get_sm__warps_active_avg_pct_of_peak_sustained_active(kernel))
            app_results.append(get_l1tex__t_sector_hit_rate_pct(kernel))
            app_results.append(get_l1tex__t_sector_pipe_lsu_mem_global_op_ld_hit_rate_pct(kernel))
            app_results.append(get_lts__t_sector_hit_rate_pct(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_ld_sum(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_st_sum(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_ld_sum(kernel) + \
                               get_l1tex__t_requests_pipe_lsu_mem_global_op_st_sum(kernel))
            app_results.append(get_l1tex__t_sectors_pipe_lsu_mem_global_op_ld_sum(kernel))
            app_results.append(get_l1tex__t_sectors_pipe_lsu_mem_global_op_st_sum(kernel))
            app_results.append(get_l1tex__t_sectors_pipe_lsu_mem_global_op_ld_sum(kernel) + \
                               get_l1tex__t_sectors_pipe_lsu_mem_global_op_st_sum(kernel))
            if get_l1tex__t_requests_pipe_lsu_mem_global_op_ld_sum(kernel) == 0.: 
                app_results.append(0.)
            else:
                app_results.append(get_l1tex__t_sectors_pipe_lsu_mem_global_op_ld_sum(kernel) / \
                                   get_l1tex__t_requests_pipe_lsu_mem_global_op_ld_sum(kernel))
            if get_l1tex__t_requests_pipe_lsu_mem_global_op_st_sum(kernel) == 0.: 
                app_results.append(0.)
            else:
                app_results.append(get_l1tex__t_sectors_pipe_lsu_mem_global_op_st_sum(kernel) / \
                                   get_l1tex__t_requests_pipe_lsu_mem_global_op_st_sum(kernel))
            app_results.append(get_l1tex__m_xbar2l1tex_read_sectors_mem_lg_op_ld_sum(kernel))
            app_results.append(get_l1tex__m_l1tex2xbar_write_sectors_mem_lg_op_st_sum(kernel))
            app_results.append(get_l1tex__m_xbar2l1tex_read_sectors_mem_lg_op_ld_sum(kernel) + \
                               get_l1tex__m_l1tex2xbar_write_sectors_mem_lg_op_st_sum(kernel))
            app_results.append(get_dram__sectors_read_sum(kernel) + get_dram__sectors_write_sum(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_atom_sum(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_red_sum(kernel))
            app_results.append(get_l1tex__t_requests_pipe_lsu_mem_global_op_atom_sum(kernel) + \
                               get_l1tex__t_requests_pipe_lsu_mem_global_op_red_sum(kernel))
            app_results.append(get_gpc__cycles_elapsed_max(kernel))
            app_results.append(get_sm__cycles_active_avg(kernel))
            app_results.append(get_smsp__inst_executed_sum(kernel))
            app_results.append(get_sm__inst_executed_avg_per_cycle_elapsed(kernel))
            app_results.append(get_sm__inst_executed_avg_per_cycle_elapsed(kernel) * \
                               get_gpc__cycles_elapsed_avg_per_second(kernel) * 1e-6)
            app_results.append(get_gpu__time_duration_sum(kernel))
            
            sheet_ncu.append(app_results + [get_L1_Total_Requests(kernel), \
                                            get_lts__t_requests_srcunit_tex_sum(kernel)])
        ################################################################################
        ####                             do ASIM report                             ####
        ################################################################################
            if not os.path.exists(report_file_path[2]+"/simulation.log"):
                result = ["NEXIST" for _ in range(36)]
                written_data_asim = [report_file_path[0].split("/")[-1].split(".")[0], str(knum)]
                for item in result[2:]:
                    written_data_asim.append(item)
                sheet_accel_sim.append(written_data_asim)
                continue

            is_deaklock = find_row_num(report_file_path[2]+"/simulation.log", \
                                       "GPGPU-Sim uArch: ERROR ** deadlock detected", 1) != -1

            kernel_info_start_line = find_row_num( \
                report_file_path[2]+"/simulation.log", "Processing kernel", knum+1)
            kernel_info_end_line = find_row_num( \
                report_file_path[2]+"/simulation.log", "gpgpu_silicon_slowdown", knum+1)
            kernel_info_start_index, kernel_info_end_index = kernel_info_start_line - 1, kernel_info_end_line - 1
            
            is_not_complete = False
            if kernel_info_start_index != -2 and kernel_info_end_index == -2:
                is_not_complete = True
            elif kernel_info_start_index == -2:
                is_not_complete = True
            
            if is_deaklock:
                result = ["DLOCK" for _ in range(34)]
            elif is_not_complete:
                result = ["NCOM" for _ in range(34)]
            else:
                result = parse_asim_result_file(report_file_path[2]+"/simulation.log", kernel_info_start_index, \
                                                kernel_info_end_index)
                if knum > 0:
                    result[13] -= last_result[13]
                    result[14] -= last_result[14]
                    result[15] -= last_result[15]
                    result[34] -= last_result[34]
                    
                last_result = copy.deepcopy(parse_asim_result_file( \
                    report_file_path[2]+"/simulation.log", kernel_info_start_index, kernel_info_end_index))
            
            written_data_asim = [report_file_path[0].split("/")[-1].split(".")[0], str(knum)]
            for item in result[2:]:
                written_data_asim.append(item)
            sheet_accel_sim.append(written_data_asim)
 
        ################################################################################
        ####                             do err report                              ####
        ################################################################################
            err_data = [report_file_path[0].split("/")[-1].split(".")[0], str(knum)]
            if is_deaklock:
                err_data += ["DLOCK" for _ in range(32)]
                err_accel_sim.append(err_data)
                continue
            elif is_not_complete:
                err_data += ["NCOM" for i in range(32)]
                err_accel_sim.append(err_data)
                continue
            else:
                
                for i in range(2, len(app_results)):
                    if i in [9, 10, 12, 13, 14, 15, 28, 29, 30, 31, 32, 33]:
                        if app_results[i] == 0:
                            err_data.append(0.)
                        else:
                            err_data.append(abs(written_data_asim[i] - app_results[i])/app_results[i])
                    else:
                        err_data.append("None")
                err_accel_sim.append(err_data)
        
        ################################################################################
        ####                             do ppt report                              ####
        ################################################################################
            result = parse_ppt_result_file(report_file_path[1]+"/kernel_"+str(knum+1)+"_SASS_g1.out")
            written_data = [report_file_path[0].split("/")[-1].split(".")[0], str(knum)]
            for item in result[2:]:
                written_data.append(item)
            sheet_ppt.append(written_data)
        
        ################################################################################
        ####                             do err report                              ####
        ################################################################################
            err_data = [report_file_path[0].split("/")[-1].split(".")[0], str(knum)]
            for i in range(2, len(app_results)):
                if app_results[i] == 0:
                    err_data.append(0.)
                else:
                    err_data.append(abs(written_data[i] - app_results[i])/app_results[i])
            err_ppt.append(err_data)
        
        ################################################################################
        ####                             do ours report                             ####
        ################################################################################
            result, result_stall = parse_ours_result_file(report_file_path[3]+"/kernel-"+str(knum)+"-summary.txt")
            
            written_data = [report_file_path[0].split("/")[-1].split(".")[0], str(knum)]
            for item in result[2:]:
                written_data.append(item)
            for item in result_stall:
                written_data.append(item)
            sheet_ours.append(written_data)
            
        ################################################################################
        ####                             do err report                              ####
        ################################################################################
            err_data = [report_file_path[0].split("/")[-1].split(".")[0], str(knum)]
            for i in range(2, len(app_results)):
                if app_results[i] == 0:
                    err_data.append(0.)
                else:
                    if isinstance(written_data[i], int) or isinstance(written_data[i], float):
                        err_data.append(abs(written_data[i] - app_results[i])/app_results[i])
                    else:
                        err_data.append("None")
            err_ours.append(err_data)

    # save
    workbook.save('compare.xlsx')
